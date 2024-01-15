# from twilio.rest import Client

# account_sid = 'AC00b0b98bdfa336cf89aaf6fb859e356f'
# auth_token = '98b6a8a67d0b9b3520211cfba36273bb'
# client = Client(account_sid, auth_token)



# message = client.messages.create(
#   from_='whatsapp:+14155238886',
#   body='Eai blz',
#   to='whatsapp:+556799021308'
# )

# print(message.sid)




import openpyxl
from datetime import datetime
from twilio.rest import Client

# Informações da conta Twilio (substitua pelos seus dados reais)
account_sid = 'AC00b0b98bdfa336cf89aaf6fb859e356f'
auth_token = '98b6a8a67d0b9b3520211cfba36273bb'
client = Client(account_sid, auth_token)

# Carregar o arquivo do estoque
workbook_estoque = openpyxl.load_workbook('estoquedez.xlsx')
sheet_estoque = workbook_estoque.active

# Criar um dicionário para mapear o código do produto ao seu estoque
# Carregar o arquivo de produtos
workbook_produtos = openpyxl.load_workbook('dados_produtos.xlsx')
sheet_produtos = workbook_produtos.active

# Verificar a validade dos produtos e criar a mensagem
estoque_produtos = {}
for row in range(2, sheet_estoque.max_row + 1):
    codigo = str(sheet_estoque.cell(row, 1).value).strip()  # Convertendo para string e removendo espaços em branco
    quantidade = sheet_estoque.cell(row, 2).value
    estoque_produtos[codigo] = quantidade


# Preparar a mensagem
mensagem = ''
hoje = datetime.now()

for row in range(2, sheet_produtos.max_row + 1):
    validade_str = sheet_produtos.cell(row, 13).value
    if validade_str:
        validade = datetime.strptime(validade_str, "%Y-%m-%d")
        if hoje < validade and (validade - hoje).days <= 30:
            codigo_produto = str(sheet_produtos.cell(row, 5).value).strip()

            if codigo_produto in estoque_produtos and estoque_produtos[codigo_produto] > 0:
                nome_produto = sheet_produtos.cell(row, 6).value
                lote = sheet_produtos.cell(row, 11).value
                nova_mensagem = f"Produto: {nome_produto}, Lote: {lote}, Validade: {validade.strftime('%d-%m-%Y')}\n"
                
                # Verificar o tamanho da mensagem
                if len(mensagem + nova_mensagem) > 1600:
                    # Enviar a mensagem atual e iniciar uma nova
                    client.messages.create(
                      from_='whatsapp:+14155238886',
                      body=mensagem,
                      to='whatsapp:+556799021308'
                    )
                    mensagem = nova_mensagem
                else:
                    mensagem += nova_mensagem

